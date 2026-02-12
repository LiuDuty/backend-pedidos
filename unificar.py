import openpyxl
import os

# Caminho do arquivo Excel
excel_path = 'src/assets/GERAL.xlsx'
output_xlsx = 'HISTORICO_UNIFICADO_COMPLETO.xlsx'

if not os.path.exists(excel_path):
    print(f"Erro: Arquivo {excel_path} não encontrado.")
    exit(1)

print(f"Carregando {excel_path} (isso pode demorar devido ao tamanho)...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

all_data = []

# Cabeçalhos solicitados pelo usuário (22 campos)
HEADERS = [
    "FORNECEDOR", "CLIENTE", "PRODUTO", "PEDIDO", "OC DO CLIENTE", "DATA DO PEDIDO", 
    "QUANT", "UND", "CAIXAS", "PREÇO POR UNIDADE", "DATA DE ENTREGA", "DATA DE FATURA", 
    "NF", "QUANT_2", "SALDO", "VALOR", "VALOR DA COMISSÃO", "C", "OBSERVAÇÃO", 
    "CÓDIGO", "PESO", "IPI"
]

for sheet_name in wb.sheetnames:
    print(f"Processando aba: {sheet_name}")
    sheet = wb[sheet_name]
    
    # Carrega as linhas da aba
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue

    # Tenta identificar o fornecedor no topo da aba (geralmente A1 ou A2)
    # Procuramos por texto que contenha o nome do fornecedor ou simplesmente o que está em A1
    potential_fornecedor = "DESCONHECIDO"
    for r_idx in range(min(5, len(rows))):
        line_text = " ".join([str(c) for c in rows[r_idx] if c])
        if "SERRAPLAST" in line_text.upper():
            potential_fornecedor = "SERRAPLAST"
            break
        if line_text.strip(): # Pega o primeiro texto significativo como fallback
            potential_fornecedor = line_text.split("  ")[0].strip()
            break

    # Encontrar as posições de "CLIENTE" na horizontal
    table_starts = []
    # Escaneamos as primeiras 15 linhas para garantir que achamos o cabeçalho
    for r_idx in range(min(15, len(rows))):
        row = rows[r_idx]
        for c_idx, cell in enumerate(row):
            if cell and str(cell).strip().upper() == "CLIENTE":
                if not any(ts['col'] == c_idx for ts in table_starts):
                    # Identificamos os índices das colunas extras se existirem no cabeçalho
                    # Procuramos por Codigo, Peso, IPI na mesma linha de cabeçalho
                    header_map = {}
                    for offset in range(30): # Procura até 30 colunas a direita do CLIENTE
                        if c_idx + offset < len(row):
                            h_cell = str(row[c_idx + offset]).strip().upper()
                            if "CÓDIGO" in h_cell or "CODIGO" in h_cell: header_map['CÓDIGO'] = offset
                            if "PESO" in h_cell: header_map['PESO'] = offset
                            if "IPI" in h_cell: header_map['IPI'] = offset
                            
                    table_starts.append({'col': c_idx, 'header_row': r_idx, 'header_map': header_map})
    
    if not table_starts:
        print(f"  ⚠️ Nenhum cabeçalho 'CLIENTE' encontrado na aba {sheet_name}.")
        continue

    for start in table_starts:
        c_start = start['col']
        r_start = start['header_row']
        h_map = start['header_map']
        
        for r in range(r_start + 1, len(rows)):
            row = rows[r]
            if c_start >= len(row) or (not row[c_start] and not row[c_start+1]):
                continue # Pula linhas vazias
            
            # Mapeia as 18 colunas padrão (A-R do bloco)
            # 0:CLIENTE, 1:PRODUTO, 2:PEDIDO, 3:OC, 4:DATA, 5:QUANT, 6:UND, 7:CAIXAS, 
            # 8:PRECO, 9:ENTREGA, 10:FATURA, 11:NF, 12:QUANT2, 13:SALDO, 14:VALOR, 15:COMISSAO, 16:C, 17:OBS
            block_18 = list(row[c_start : c_start + 18])
            while len(block_18) < 18: block_18.append(None)
            
            # Valor do Fornecedor é a primeira coluna
            # Colunas extras (Código, Peso, IPI)
            codigo = row[c_start + h_map['CÓDIGO']] if 'CÓDIGO' in h_map and c_start + h_map['CÓDIGO'] < len(row) else None
            peso = row[c_start + h_map['PESO']] if 'PESO' in h_map and c_start + h_map['PESO'] < len(row) else None
            ipi = row[c_start + h_map['IPI']] if 'IPI' in h_map and c_start + h_map['IPI'] < len(row) else None
            
            final_row = [potential_fornecedor] + block_18 + [codigo, peso, ipi]
            all_data.append(final_row)

# Gerar XLSX Final
print(f"\nGerando Excel com {len(all_data)} linhas...")
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = "Histórico Unificado"
new_sheet.append(HEADERS + ["ABA_ORIGINAL"])

for line in all_data:
    new_sheet.append(line)

new_wb.save(output_xlsx)
print(f"✅ Sucesso! Arquivo gerado: {output_xlsx}")