import openpyxl
import sqlite3
import os
from datetime import datetime

DB_PATH = 'notas.db'
EXCEL_PATH = 'src/assets/NOTAS.xlsx'
HISTORICO_PATH = 'HISTORICO_UNIFICADO_COMPLETO.xlsx'

def connect_db():
    return sqlite3.connect(DB_PATH)

def import_suppliers(cursor, wb):
    print("Importando Fornecedores...")
    sheet = wb['Fornecedores']
    # Cabeçalho na Linha 1: Identificação, Logo, Fornecedor, ENDEREÇO, Nº, CEP, ...
    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(row=r, column=3).value
        if not name: continue
        
        logo = sheet.cell(row=r, column=2).value
        address = sheet.cell(row=r, column=4).value
        number = sheet.cell(row=r, column=5).value
        zipcode = sheet.cell(row=r, column=6).value
        
        cursor.execute('''
            INSERT OR IGNORE INTO suppliers (name, logo_filename, address, number, zipcode, createdAt, updatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (str(name).strip(), str(logo).strip() if logo else '', str(address).strip() if address else '', 
              str(number).strip() if number else '', str(zipcode).strip() if zipcode else '', 
              datetime.now(), datetime.now()))

def import_customers(cursor, wb):
    print("Importando Clientes...")
    sheet = wb['Clientes']
    # Cabeçalho na Linha 1: Identificação, CLIENTE, ENDEREÇO, Nº, CEP, CAIXA POSTAL, BAIRRO, CIDADE, ESTADO, FONE, CONTATO, E-MAIL, CNPJ, IE
    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(row=r, column=2).value
        if not name: continue
        
        address = sheet.cell(row=r, column=3).value
        number = sheet.cell(row=r, column=4).value
        zipcode = sheet.cell(row=r, column=5).value
        neighborhood = sheet.cell(row=r, column=7).value
        city = sheet.cell(row=r, column=8).value
        state = sheet.cell(row=r, column=9).value
        phone = sheet.cell(row=r, column=10).value
        contact = sheet.cell(row=r, column=11).value
        email = sheet.cell(row=r, column=12).value
        cnpj = sheet.cell(row=r, column=13).value
        ie = sheet.cell(row=r, column=14).value

        cursor.execute('''
            INSERT OR IGNORE INTO customers (name, address, number, zipcode, neighborhood, city, state, phone, contact, email, cnpj, state_registration, createdAt, updatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (str(name).strip(), str(address).strip() if address else '', str(number).strip() if number else '',
              str(zipcode).strip() if zipcode else '', str(neighborhood).strip() if neighborhood else '',
              str(city).strip() if city else '', str(state).strip() if state else '',
              str(phone).strip() if phone else '', str(contact).strip() if contact else '',
              str(email).strip() if email else '', str(cnpj).strip() if cnpj else '',
              str(ie).strip() if ie else '', datetime.now(), datetime.now()))

def import_history(cursor, wb_notas=None):
    source_path = HISTORICO_PATH
    use_notas_sheet = False
    
    if os.path.exists(HISTORICO_PATH):
        print(f"Importando Histórico de {HISTORICO_PATH}...")
        wb = openpyxl.load_workbook(HISTORICO_PATH, data_only=True)
        sheet = wb.active
        # Mapeamento HISTORICO_UNIFICADO_COMPLETO.xlsx: 
        # Col 0:Data, 1:Pedido, 2:Fornecedor, 3:CLIENTE, 4:PRODUTO, 6:QUANT, 8:PRECO
        col_map = {'date': 1, 'order': 2, 'supplier': 3, 'customer': 4, 'prod': 5, 'qty': 7, 'price': 9}
    elif wb_notas and 'Historico' in wb_notas.sheetnames:
        print("Importando Histórico da aba 'Historico' do NOTAS.xlsx...")
        sheet = wb_notas['Historico']
        use_notas_sheet = True
        # Mapeamento NOTAS.xlsx['Historico']: 
        # Col 0:FORNECEDOR, 1:CLIENTE, 2:PEDIDO, 3:OC, 4:DATA, 5:PRODUTO, 6:UND, 7:QUANT, 8:PREÇO
        col_map = {'supplier': 1, 'customer': 2, 'order': 3, 'date': 5, 'prod': 6, 'qty': 8, 'price': 9}
    else:
        print("Nenhuma fonte de histórico encontrada.")
        return

    # Iniciar mapeamento de IDs (Case-insensitive)
    cursor.execute("SELECT id, name FROM suppliers")
    suppliers_map = {name.strip().upper(): id for id, name in cursor.fetchall()}
    cursor.execute("SELECT id, name FROM customers")
    customers_map = {name.strip().upper(): id for id, name in cursor.fetchall()}

    count = 0
    max_r = sheet.max_row
    
    last_s_id = None
    last_c_id = None
    last_order_num = None
    last_date_str = None

    for r in range(2, max_r + 1):
        def get_val(key): 
            try: return sheet.cell(row=r, column=col_map[key]).value
            except: return None
        
        date_val = get_val('date')
        order_num = get_val('order')
        supplier_name = get_val('supplier')
        customer_name = get_val('customer')
        
        # Fill-down logic
        if not supplier_name and last_s_id:
            s_id = last_s_id
        elif supplier_name:
            s_name_clean = str(supplier_name).strip().upper()
            if s_name_clean not in suppliers_map:
                cursor.execute("INSERT INTO suppliers (name, createdAt, updatedAt) VALUES (?, ?, ?)", 
                              (str(supplier_name).strip(), datetime.now(), datetime.now()))
                s_id = cursor.lastrowid
                suppliers_map[s_name_clean] = s_id
            else:
                s_id = suppliers_map[s_name_clean]
            last_s_id = s_id
        else:
            continue

        if not customer_name and last_c_id:
            c_id = last_c_id
        elif customer_name:
            c_name_clean = str(customer_name).strip().upper()
            if c_name_clean not in customers_map:
                cursor.execute("INSERT INTO customers (name, createdAt, updatedAt) VALUES (?, ?, ?)", 
                              (str(customer_name).strip(), datetime.now(), datetime.now()))
                c_id = cursor.lastrowid
                customers_map[c_name_clean] = c_id
            else:
                c_id = customers_map[c_name_clean]
            last_c_id = c_id
        else:
            continue

        if not order_num and last_order_num:
            current_order_num = last_order_num
        elif order_num:
            current_order_num = str(order_num).strip()
            last_order_num = current_order_num
        else:
            continue

        # Formatar data com fill-down
        date_str = None
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%Y-%m-%d')
        elif date_val:
            d_str = str(date_val).strip()
            if '/' in d_str:
                try: 
                    parts = d_str.split('/')
                    if len(parts) == 3:
                        date_str = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                except: pass
            elif '-' in d_str and len(d_str) >= 10:
                date_str = d_str[:10]
        
        if not date_str:
            date_str = last_date_str if last_date_str else datetime.now().strftime('%Y-%m-%d')
        
        last_date_str = date_str

        try:
            cursor.execute('''
                INSERT OR IGNORE INTO orders (orderNumber, orderDate, supplierId, customerId, createdAt, updatedAt)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (current_order_num, date_str, s_id, c_id, datetime.now(), datetime.now()))
            
            cursor.execute("SELECT id FROM orders WHERE orderNumber = ? AND supplierId = ? AND customerId = ?", 
                          (current_order_num, s_id, c_id))
            row = cursor.fetchone()
            if not row: continue
            order_id = row[0]
            
            prod_name = str(get_val('prod') or "Produto Histórico").strip()
            qty = get_val('qty') or 0
            price = get_val('price') or 0
            
            # Verificar se ítem já existe
            cursor.execute("SELECT id FROM orderItems WHERE orderId = ? AND productName = ? AND quantity = ?", 
                          (order_id, prod_name, qty))
            if not cursor.fetchone():
                cursor.execute('''
                    INSERT INTO orderItems (productName, quantity, pricePerThousand, orderId, createdAt, updatedAt)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (prod_name, qty, price, order_id, datetime.now(), datetime.now()))
                count += 1
        except Exception:
            pass

    print(f"Importados {count} registros de itens de pedido.")


def main():
    if not os.path.exists(DB_PATH):
        print(f"Erro: Banco de dados {DB_PATH} não encontrado.")
        return

    conn = connect_db()
    cursor = conn.cursor()
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    import_suppliers(cursor, wb)
    import_customers(cursor, wb)
    conn.commit()
    
    import_history(cursor, wb)
    
    conn.commit()
    conn.close()
    print("Processo concluído!")

if __name__ == "__main__":
    main()
