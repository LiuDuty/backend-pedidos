import openpyxl

def diagnose():
    try:
        wb = openpyxl.load_workbook('src/assets/NOTAS.xlsx', data_only=True)
        if 'Historico' not in wb.sheetnames:
            print("Aba 'Historico' não encontrada.")
            return
            
        sheet = wb['Historico']
        total = sheet.max_row - 1
        m_p = 0
        m_s = 0
        m_c = 0
        v = 0
        
        for r in range(2, sheet.max_row + 1):
            s = sheet.cell(row=r, column=1).value
            c = sheet.cell(row=r, column=2).value
            p = sheet.cell(row=r, column=3).value
            
            if not p: m_p += 1
            if not s: m_s += 1
            if not c: m_c += 1
            
            if p and s and c: v += 1
            
        print(f"DEBUG_RESULT:Total={total},MissingP={m_p},MissingS={m_s},MissingC={m_c},Valid={v}")
    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    diagnose()
